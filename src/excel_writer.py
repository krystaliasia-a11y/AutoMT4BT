from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from src.report_parser import BacktestResult

HEADERS = [
    "Set檔名",
    "EA",
    "幣對",
    "週期",
    "日期範圍",
    "初始資金",
    "淨利潤",
    "Profit Factor",
    "最大回撤",
    "最大回撤%",
    "總交易數",
    "勝率%",
    "期望收益",
    "模擬品質",
]


def write_results(results: list[BacktestResult], output_path: str) -> str:
    """將所有回測結果寫入 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "回測結果"

    # 標題列樣式
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # 寫入資料
    for row_idx, r in enumerate(results, 2):
        data = [
            r.set_filename,
            r.expert,
            r.symbol,
            r.period,
            r.date_range,
            r.deposit,
            r.total_net_profit,
            r.profit_factor,
            r.maximal_drawdown,
            r.maximal_drawdown_pct,
            r.total_trades,
            r.win_rate,
            r.expected_payoff,
            r.modeling_quality,
        ]
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # 淨利潤為負值標紅
            if col_idx == 7 and isinstance(value, (int, float)) and value < 0:
                cell.font = Font(color="FF0000")

        # 數字格式
        ws.cell(row=row_idx, column=6).number_format = "#,##0"
        ws.cell(row=row_idx, column=7).number_format = "#,##0.00"
        ws.cell(row=row_idx, column=8).number_format = "0.00"
        ws.cell(row=row_idx, column=9).number_format = "#,##0.00"
        ws.cell(row=row_idx, column=10).number_format = "0.00\"%\""
        ws.cell(row=row_idx, column=12).number_format = "0.00\"%\""
        ws.cell(row=row_idx, column=13).number_format = "0.00"

    # 自動調整欄寬
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    # 凍結標題列
    ws.freeze_panes = "A2"

    wb.save(output_path)
    return output_path
